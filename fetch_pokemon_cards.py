#!/usr/bin/env python3
"""
Fetch all Pokémon TCG cards set-by-set to avoid deep-pagination rate limits.
Instead of paginating globally (which triggers a 400 at ~page 54), we:
  1. Fetch the full list of sets from /v2/sets  (1-2 pages, fast)
  2. For each set, fetch cards with ?q=set.id:{id}  (most sets fit in 1 page)
  3. Merge, sort, and write to Excel

Usage:
  python3 fetch_pokemon_cards.py [output.xlsx] [--resume sets_done.txt]

Options:
  output.xlsx        Path for the output file (default: pokemon_cards_pricechart.xlsx)
  --resume FILE      Resume from a checkpoint file listing already-fetched set IDs
                     (checkpoint is written automatically as sets are completed)
"""

import sys
import os
import re
import time
import json
import requests
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── helpers ──────────────────────────────────────────────────────────────────

def make_slug(name):
    if not name:
        return ""
    slug = name.lower()
    slug = re.sub(r'[^a-z0-9\s\-]', '', slug)
    slug = re.sub(r'\s+', '-', slug.strip())
    slug = re.sub(r'-+', '-', slug)
    return slug.strip('-')

def get_with_retry(url, params=None, max_retries=6, label=""):
    """GET a URL with exponential back-off on 429 / 400 / network errors."""
    for attempt in range(max_retries):
        try:
            r = requests.get(url, params=params, timeout=30)
            if r.status_code in (429, 400):
                wait = 15 * (attempt + 1)
                print(f"  [{label}] Rate-limited ({r.status_code}), attempt {attempt+1}/{max_retries}. Waiting {wait}s…")
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json()
        except requests.exceptions.RequestException as e:
            wait = 10 * (attempt + 1)
            print(f"  [{label}] Error attempt {attempt+1}/{max_retries}: {e}. Waiting {wait}s…")
            time.sleep(wait)
    return None

# ── step 1: fetch all sets ────────────────────────────────────────────────────

def fetch_all_sets():
    sets = []
    page = 1
    print("Fetching set list…")
    while True:
        data = get_with_retry(
            "https://api.pokemontcg.io/v2/sets",
            params={"pageSize": 250, "page": page},
            label=f"sets p{page}"
        )
        if not data:
            print("  Failed to fetch sets page, stopping.")
            break
        batch = data.get("data", [])
        if not batch:
            break
        sets.extend(batch)
        print(f"  Sets page {page}: {len(batch)} sets (total {len(sets)})")
        if len(sets) >= data.get("totalCount", 0):
            break
        page += 1
        time.sleep(0.5)
    print(f"Total sets: {len(sets)}")
    return sets

# ── step 2: fetch cards for one set ──────────────────────────────────────────

def fetch_cards_for_set(set_id, set_name):
    cards = []
    page = 1
    while True:
        data = get_with_retry(
            "https://api.pokemontcg.io/v2/cards",
            params={
                "q": f"set.id:{set_id}",
                "pageSize": 250,
                "page": page,
            },
            label=f"{set_id} p{page}"
        )
        if not data:
            print(f"  [{set_id}] Failed to fetch page {page}, stopping.")
            break
        batch = data.get("data", [])
        if not batch:
            break
        cards.extend(batch)
        total = data.get("totalCount", 0)
        if page == 1:
            print(f"  {set_id} ({set_name}): {total} cards", end="", flush=True)
        if len(cards) >= total:
            print(f" ✓")
            break
        print(f" p{page}", end="", flush=True)
        page += 1
        time.sleep(0.6)
    return cards

# ── step 3: build Excel ───────────────────────────────────────────────────────

def card_to_row(card):
    name        = card.get("name", "")
    set_info    = card.get("set", {})
    set_name    = set_info.get("name", "")
    set_code    = set_info.get("id", "")
    card_number = card.get("number", "")
    rarity      = card.get("rarity", "")
    supertype   = card.get("supertype", "")

    tcgplayer_url = card.get("tcgplayer", {}).get("url", "")
    image_url     = card.get("images", {}).get("small", "")

    slug = make_slug(name)
    pricecharting_url = f"https://www.pricecharting.com/game/pokemon/{slug}"
    search_url        = f"https://www.pricecharting.com/search-products?q={quote(name)}&type=prices"

    return [name, set_name, set_code, card_number, rarity, supertype,
            pricecharting_url, search_url, tcgplayer_url, image_url]

def create_excel(rows, output_path):
    print(f"\nWriting {len(rows)} rows to {output_path}…")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pokemon Cards"

    headers = [
        "Card Name", "Set Name", "Set Code", "Card Number",
        "Rarity", "Supertype",
        "PriceCharting URL", "PriceCharting Search URL",
        "TCGPlayer URL", "Card Image URL",
    ]

    header_font  = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_fill  = PatternFill(start_color='2D2D2D', end_color='2D2D2D', fill_type='solid')
    data_font    = Font(name='Arial', size=9)
    white_fill   = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    gray_fill    = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 20

    for row_idx, row_data in enumerate(rows, 2):
        fill = gray_fill if row_idx % 2 == 0 else white_fill
        ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font      = data_font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

    widths = [30, 28, 10, 8, 18, 12, 55, 55, 55, 55]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}  ({len(rows)} cards)")

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    output_path    = "pokemon_cards_pricechart.xlsx"
    resume_file    = None
    checkpoint_file = "fetch_checkpoint.txt"   # written automatically

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--resume":
            resume_file = args[i + 1]
            i += 2
        else:
            output_path = args[i]
            i += 1

    # Load checkpoint (sets already completed)
    done_sets = set()
    cf = resume_file or checkpoint_file
    if os.path.exists(cf):
        with open(cf) as f:
            done_sets = {line.strip() for line in f if line.strip()}
        print(f"Resuming — {len(done_sets)} sets already fetched (from {cf})")

    sets = fetch_all_sets()
    if not sets:
        print("No sets fetched. Check your internet connection.")
        sys.exit(1)

    all_rows = []
    total_sets = len(sets)

    for idx, s in enumerate(sets, 1):
        set_id   = s["id"]
        set_name = s["name"]

        if set_id in done_sets:
            print(f"[{idx}/{total_sets}] Skipping {set_id} (already done)")
            continue

        print(f"[{idx}/{total_sets}] ", end="", flush=True)
        cards = fetch_cards_for_set(set_id, set_name)

        for card in cards:
            all_rows.append(card_to_row(card))

        # Mark this set done in the checkpoint file
        with open(checkpoint_file, "a") as f:
            f.write(set_id + "\n")
        done_sets.add(set_id)

        # Polite pause between sets
        time.sleep(0.8)

    # Sort by set_code then card_number
    all_rows.sort(key=lambda r: (r[2], r[3] if r[3] else ""))

    create_excel(all_rows, output_path)

    # Clean up checkpoint on success
    if os.path.exists(checkpoint_file):
        os.remove(checkpoint_file)
        print("Checkpoint file removed (fetch complete).")

if __name__ == "__main__":
    main()
