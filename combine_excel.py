#!/usr/bin/env python3
"""
Combine two pokemon card Excel files into one.
Usage: python3 combine_excel.py file1.xlsx file2.xlsx combined_output.xlsx
"""

import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def combine(file1, file2, output):
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    ws1 = wb1.active
    ws2 = wb2.active

    # Headers from each file
    headers1 = [cell.value for cell in ws1[1]]
    headers2 = [cell.value for cell in ws2[1]]

    # Output uses file 1's headers (preserves any extra columns like Price Source URL)
    # Any columns in file 2 not in file 1 are appended at the end
    out_headers = list(headers1)
    for h in headers2:
        if h not in out_headers:
            out_headers.append(h)
    n_cols = len(out_headers)

    # Build column index maps: header name -> output column index
    col_map1 = {h: i for i, h in enumerate(headers1)}
    col_map2 = {h: i for i, h in enumerate(headers2)}
    out_idx  = {h: i for i, h in enumerate(out_headers)}

    # Find set_code and card_number column indices in the output (for dedup key)
    set_code_idx   = out_idx.get('Set Code',    out_idx.get('set_code',   3))
    card_num_idx   = out_idx.get('Card Number', out_idx.get('card_number', 4))

    def align_row(raw_row, col_map):
        """Map a raw row tuple into the unified output column order."""
        out = [None] * n_cols
        for header, src_idx in col_map.items():
            if src_idx < len(raw_row):
                dst_idx = out_idx.get(header)
                if dst_idx is not None:
                    out[dst_idx] = raw_row[src_idx]
        return out

    # Collect and align rows from each file
    rows_f1 = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows_f1.append(align_row(row, col_map1))

    rows_f2 = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows_f2.append(align_row(row, col_map2))

    def norm_card_num(val):
        """Normalize card number to a clean string: 1.0 -> '1', '25a' -> '25a'"""
        if val is None:
            return ''
        s = str(val).strip()
        # Strip trailing '.0' from floats that Excel stores as numbers
        if s.endswith('.0') and s[:-2].isdigit():
            s = s[:-2]
        return s

    def card_key(row):
        sc = str(row[set_code_idx] or '').strip()
        cn = norm_card_num(row[card_num_idx])
        return (sc, cn)

    def sort_key(row):
        sc = str(row[set_code_idx] or '')
        cn = norm_card_num(row[card_num_idx])
        # natural sort: numeric prefix as int, then full string for ties (e.g. 25a vs 25b)
        num = ''
        for ch in cn:
            if ch.isdigit():
                num += ch
            else:
                break
        return (sc, int(num) if num else 0, cn)

    # Dedup: file 1 rows take priority (preserving Price Source URL and any other extras)
    # Also normalize card number back to clean string while we're at it
    seen = set()
    unique_rows = []
    for row in rows_f1:
        # Fix float card numbers in-place
        row[card_num_idx] = norm_card_num(row[card_num_idx])
        key = card_key(row)
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)

    new_from_f2 = 0
    for row in rows_f2:
        key = card_key(row)
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)
            new_from_f2 += 1

    unique_rows.sort(key=sort_key)

    print(f"File 1: {len(rows_f1)} data rows  (headers: {headers1})")
    print(f"File 2: {len(rows_f2)} data rows  (headers: {headers2})")
    print(f"Output columns: {out_headers}")
    print(f"New cards added from File 2: {new_from_f2}")
    print(f"Combined (deduplicated): {len(unique_rows)} rows")

    # Write output
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Pokemon Cards"

    header_font  = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_fill  = PatternFill(start_color='2D2D2D', end_color='2D2D2D', fill_type='solid')
    data_font    = Font(name='Arial', size=9)
    white_fill   = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    gray_fill    = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    # Write header
    out_ws.append(out_headers)
    for col_idx, _ in enumerate(out_headers, 1):
        cell = out_ws.cell(row=1, column=col_idx)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
    out_ws.row_dimensions[1].height = 20

    # Write data
    for row_idx, row_data in enumerate(unique_rows, 2):
        fill = gray_fill if row_idx % 2 == 0 else white_fill
        out_ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = out_ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='top')

    out_ws.freeze_panes = "A2"
    out_wb.save(output)
    print(f"Saved to: {output}")

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python3 combine_excel.py file1.xlsx file2.xlsx output.xlsx")
        sys.exit(1)
    combine(sys.argv[1], sys.argv[2], sys.argv[3])
